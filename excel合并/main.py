from openpyxl import load_workbook
from openpyxl import Workbook

# 加载现有的xlsx文件
read_excel_first = load_workbook(filename='first.xlsx')
read_excel_second = load_workbook(filename='second.xlsx')


# 激活工作表  
r_excel_first = read_excel_first.active
r_excel_second = read_excel_second.active

# 创建一个工作簿
write_excel = Workbook()

# 激活工作表
w_excel = write_excel.active

# 标头信息
w_excel['A1'] = "设备唯一编码"
w_excel['A2'] = "field_code"
w_excel['B1'] = "设备类型"
w_excel['B2'] = "channel_type_code"
w_excel['C1'] = "设备中文名称"
w_excel['C2'] = "name"
w_excel['D1'] = "指标中文（用于自己比对的，无实际作用）"
w_excel['D2'] = "kpi_name"
w_excel['E1'] = "指标英文(需要去后勤4.0 指标体系.xls中根据指标中文查找对应的)"
w_excel['E2'] = "kpi_code"
w_excel['F1'] = "读写类型（0 只读 1 读写 2 写）"
w_excel['F2'] = "read_write_type"
w_excel['G1'] = "采集地址（第一位是功能码，后四位是实际地址）(值）"
w_excel['G2'] = "address"
w_excel['H1'] = "位数"
w_excel['H2'] = "length"
w_excel['I1'] = "系数（默认 1）"
w_excel['I2'] = "factor"
w_excel['J1'] = "网关地址（IP）"
w_excel['J2'] = "network_ip"
w_excel['K1'] = "网关地址（端口）"
w_excel['K2'] = "network_port"
w_excel['L1'] = "数值类型（0 整形 1 浮点型）"
w_excel['L2'] = "data_type"
w_excel['M1'] = "存储类型(0线圈/1寄存器)"
w_excel['M2'] = "save_type"
w_excel['N1'] = "对应模块"
w_excel['N2'] = "module"
w_excel['O1'] = "版本"
w_excel['O2'] = "version"
w_excel['P1'] = "回路序号"
w_excel['P2'] = "circuit_number"
w_excel['Q1'] = "回路编号"
w_excel['Q2'] = "circuit_id"
w_excel['R1'] = "备注"
w_excel['R2'] = "remarks"

# 跳过标头数据
second_rows_temp = r_excel_second.iter_rows(values_only=True)
next(second_rows_temp)
first_rows_temp = r_excel_first.iter_rows(values_only=True)
next(first_rows_temp)

first_rows = []
second_rows = []

for row in first_rows_temp:
    first_rows.append(row)
for row in second_rows_temp:
    second_rows.append(row)

temp_name = ""
temp_num = 0
first_dict = {}

# 获取first.xlsx的字典集合
for row in first_rows:
    if row[0] != temp_name:
        temp_name = row[0]
        temp_num = 1
    else:
        temp_num += 1
    row = (*row,temp_num,row[0]+"_"+str(temp_num))
    first_dict[row[-1]] = row

# 整理second_rows数据
second_rows_temp = second_rows
second_rows = []
temp_rows = []
for row in second_rows_temp:
    if "心跳" in row[2]:
        # print(temp_rows)
        num = 1
        for temp_row in temp_rows:
            temp_row = temp_row[-1]
            if first_dict[temp_row][3] == num:
                num += 1
                temp = (row[0],row[1],"kpi_a01_a001",first_dict[temp_row][-1])
                second_rows.append(temp)
            else:
                temp_rows = temp_rows[num-1:]
                break
        else:
            temp_rows = temp_rows[num-1:]
    elif "单控" in row[2]:
        temp_row = (row[0],row[1],"kpi_a01_a002",row[2].replace("单控","_"))
        second_rows.append(temp_row)
        temp_rows.append(temp_row)

num = 3
# 整合输出数据，头2行是表头
for row in second_rows:
    name = row[-1]
    w_excel['A'+str(num)] = name
    w_excel['C'+str(num)] = first_dict[name][5]
    w_excel['E'+str(num)] = row[2]
    type = ""
    if row[1] == "开关读":
        type = "0"
    elif row[1] == "开关读写":
        type = "1"
    w_excel['F'+str(num)] = type
    w_excel['G'+str(num)] = row[0]
    w_excel['H'+str(num)] = 1
    w_excel['I'+str(num)] = 1
    w_excel['J'+str(num)] = first_dict[name][6]
    w_excel['K'+str(num)] = 502
    w_excel['L'+str(num)] = 1
    w_excel['M'+str(num)] = 1
    w_excel['N'+str(num)] = first_dict[name][1]
    w_excel['O'+str(num)] = first_dict[name][2]
    w_excel['P'+str(num)] = first_dict[name][3]
    w_excel['Q'+str(num)] = first_dict[name][4]

    num += 1

# 保存工作簿
write_excel.save("result.xlsx")